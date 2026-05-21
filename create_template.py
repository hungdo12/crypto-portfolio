"""Tạo template Excel/Google Sheets cho quản lý đầu tư crypto."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.table import Table, TableStyleInfo

wb = Workbook()

# Common styles
HEADER_FILL = PatternFill('solid', start_color='1F2937')
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
INPUT_FONT = Font(name='Arial', color='0000FF', size=10)
FORMULA_FONT = Font(name='Arial', color='000000', size=10)
LINK_FONT = Font(name='Arial', color='008000', size=10)
KEY_FILL = PatternFill('solid', start_color='FFFF00')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
RIGHT = Alignment(horizontal='right', vertical='center')
THIN_BORDER = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB'),
)

# =================== SHEET 1: TRANSACTIONS (Lịch sử giao dịch) ===================
ws_tx = wb.active
ws_tx.title = 'Transactions'

tx_headers = [
    'Ngày', 'Loại', 'Token', 'Mạng', 'Contract Address',
    'Số lượng Token', 'Giá/Token (USD)', 'Tổng giá trị (USD)',
    'Đồng thanh toán', 'Phí Gas (USD)', 'Tx Hash', 'Ghi chú'
]

for col, header in enumerate(tx_headers, 1):
    cell = ws_tx.cell(row=1, column=col, value=header)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = THIN_BORDER

# Sample data
sample_tx = [
    ['2025-01-15', 'BUY', 'PEPE', 'ETH', '0x6982508145454ce325ddbe47a25d4ec3d2311933',
     1000000, 0.0000085, '=F2*G2', 'USDT', 5.20, '0xabc...', 'DCA lần 1'],
    ['2025-02-10', 'BUY', 'PEPE', 'ETH', '0x6982508145454ce325ddbe47a25d4ec3d2311933',
     500000, 0.0000095, '=F3*G3', 'USDT', 4.80, '0xdef...', 'DCA lần 2'],
    ['2025-03-05', 'BUY', 'BRETT', 'BASE', '0x532f27101965dd16442e59d40670faf5ebb142e4',
     5000, 0.085, '=F4*G4', 'USDT', 0.50, '0xghi...', 'Mua mới'],
]

for row_idx, row_data in enumerate(sample_tx, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_tx.cell(row=row_idx, column=col_idx, value=value)
        cell.border = THIN_BORDER
        if col_idx == 8:  # Tổng giá trị = formula
            cell.font = FORMULA_FONT
        else:
            cell.font = INPUT_FONT

# Format cột
ws_tx.column_dimensions['A'].width = 12
ws_tx.column_dimensions['B'].width = 8
ws_tx.column_dimensions['C'].width = 10
ws_tx.column_dimensions['D'].width = 8
ws_tx.column_dimensions['E'].width = 45
ws_tx.column_dimensions['F'].width = 15
ws_tx.column_dimensions['G'].width = 15
ws_tx.column_dimensions['H'].width = 15
ws_tx.column_dimensions['I'].width = 12
ws_tx.column_dimensions['J'].width = 12
ws_tx.column_dimensions['K'].width = 20
ws_tx.column_dimensions['L'].width = 25

# Số định dạng
for row in range(2, 100):
    ws_tx.cell(row=row, column=6).number_format = '#,##0.########'
    ws_tx.cell(row=row, column=7).number_format = '$#,##0.00000000'
    ws_tx.cell(row=row, column=8).number_format = '$#,##0.00'
    ws_tx.cell(row=row, column=10).number_format = '$#,##0.00'

# Freeze hàng đầu
ws_tx.freeze_panes = 'A2'

# =================== SHEET 2: PORTFOLIO (Tổng hợp danh mục) ===================
ws_pf = wb.create_sheet('Portfolio')

pf_headers = [
    'Token', 'Mạng', 'Contract', 'Tổng SL Mua', 'Tổng SL Bán', 'SL Đang Giữ',
    'Tổng Chi USD', 'Giá TB Mua', 'Giá Hiện Tại', 'Giá Trị Hiện Tại',
    'P&L USD', 'P&L %', 'Target Chốt Lời', 'Stop Loss', '% Danh Mục'
]

for col, header in enumerate(pf_headers, 1):
    cell = ws_pf.cell(row=1, column=col, value=header)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = THIN_BORDER

# Sample tokens trong portfolio (auto-summary từ Transactions)
portfolio_tokens = [
    ('PEPE', 'ETH', '0x6982508145454ce325ddbe47a25d4ec3d2311933'),
    ('BRETT', 'BASE', '0x532f27101965dd16442e59d40670faf5ebb142e4'),
]

for row_idx, (token, network, contract) in enumerate(portfolio_tokens, 2):
    # A: Token (input)
    ws_pf.cell(row=row_idx, column=1, value=token).font = INPUT_FONT
    # B: Mạng (input)
    ws_pf.cell(row=row_idx, column=2, value=network).font = INPUT_FONT
    # C: Contract (input)
    ws_pf.cell(row=row_idx, column=3, value=contract).font = INPUT_FONT
    # D: Tổng SL Mua = SUMIFS từ Transactions
    ws_pf.cell(row=row_idx, column=4,
               value=f'=SUMIFS(Transactions!F:F,Transactions!C:C,A{row_idx},Transactions!B:B,"BUY")').font = LINK_FONT
    # E: Tổng SL Bán
    ws_pf.cell(row=row_idx, column=5,
               value=f'=SUMIFS(Transactions!F:F,Transactions!C:C,A{row_idx},Transactions!B:B,"SELL")').font = LINK_FONT
    # F: SL Đang Giữ = D - E
    ws_pf.cell(row=row_idx, column=6, value=f'=D{row_idx}-E{row_idx}').font = FORMULA_FONT
    # G: Tổng Chi USD = SUMIFS giá trị BUY
    ws_pf.cell(row=row_idx, column=7,
               value=f'=SUMIFS(Transactions!H:H,Transactions!C:C,A{row_idx},Transactions!B:B,"BUY")+SUMIFS(Transactions!J:J,Transactions!C:C,A{row_idx})').font = LINK_FONT
    # H: Giá TB Mua = G / D (weighted average, có phòng chia 0)
    ws_pf.cell(row=row_idx, column=8,
               value=f'=IFERROR(G{row_idx}/D{row_idx},0)').font = FORMULA_FONT
    # I: Giá Hiện Tại (do bot Python cập nhật - YELLOW)
    cell_i = ws_pf.cell(row=row_idx, column=9, value=0.0000095 if token == 'PEPE' else 0.092)
    cell_i.font = INPUT_FONT
    cell_i.fill = KEY_FILL
    # J: Giá Trị Hiện Tại = F * I
    ws_pf.cell(row=row_idx, column=10, value=f'=F{row_idx}*I{row_idx}').font = FORMULA_FONT
    # K: P&L USD = J - G
    ws_pf.cell(row=row_idx, column=11, value=f'=J{row_idx}-G{row_idx}').font = FORMULA_FONT
    # L: P&L % = (J - G) / G
    ws_pf.cell(row=row_idx, column=12,
               value=f'=IFERROR((J{row_idx}-G{row_idx})/G{row_idx},0)').font = FORMULA_FONT
    # M: Target chốt lời (input)
    ws_pf.cell(row=row_idx, column=13, value=0.00005 if token == 'PEPE' else 0.5).font = INPUT_FONT
    # N: Stop loss (input)
    ws_pf.cell(row=row_idx, column=14, value=0.000005 if token == 'PEPE' else 0.05).font = INPUT_FONT
    # O: % danh mục = J / TOTAL
    ws_pf.cell(row=row_idx, column=15,
               value=f'=IFERROR(J{row_idx}/SUM($J$2:$J$50),0)').font = FORMULA_FONT

# Hàng tổng (Total)
total_row = len(portfolio_tokens) + 3
ws_pf.cell(row=total_row, column=1, value='TỔNG').font = Font(bold=True, size=11)
ws_pf.cell(row=total_row, column=7, value=f'=SUM(G2:G{total_row-2})').font = Font(bold=True, color='000000')
ws_pf.cell(row=total_row, column=10, value=f'=SUM(J2:J{total_row-2})').font = Font(bold=True, color='000000')
ws_pf.cell(row=total_row, column=11, value=f'=SUM(K2:K{total_row-2})').font = Font(bold=True, color='000000')
ws_pf.cell(row=total_row, column=12, value=f'=IFERROR(K{total_row}/G{total_row},0)').font = Font(bold=True, color='000000')

for col in range(1, 16):
    ws_pf.cell(row=total_row, column=col).fill = PatternFill('solid', start_color='E5E7EB')

# Column widths
widths = [10, 8, 45, 15, 15, 15, 15, 15, 15, 17, 15, 10, 15, 15, 12]
for i, w in enumerate(widths, 1):
    ws_pf.column_dimensions[get_column_letter(i)].width = w

# Number formats
for row in range(2, 50):
    ws_pf.cell(row=row, column=4).number_format = '#,##0.########'
    ws_pf.cell(row=row, column=5).number_format = '#,##0.########'
    ws_pf.cell(row=row, column=6).number_format = '#,##0.########'
    ws_pf.cell(row=row, column=7).number_format = '$#,##0.00'
    ws_pf.cell(row=row, column=8).number_format = '$#,##0.00000000'
    ws_pf.cell(row=row, column=9).number_format = '$#,##0.00000000'
    ws_pf.cell(row=row, column=10).number_format = '$#,##0.00'
    ws_pf.cell(row=row, column=11).number_format = '$#,##0.00;[Red]($#,##0.00)'
    ws_pf.cell(row=row, column=12).number_format = '0.00%;[Red](0.00%)'
    ws_pf.cell(row=row, column=13).number_format = '$#,##0.00000000'
    ws_pf.cell(row=row, column=14).number_format = '$#,##0.00000000'
    ws_pf.cell(row=row, column=15).number_format = '0.00%'

# Conditional formatting cho P&L
green_fill = PatternFill('solid', start_color='D1FAE5')
red_fill = PatternFill('solid', start_color='FEE2E2')
ws_pf.conditional_formatting.add(f'K2:K{total_row-2}',
    CellIsRule(operator='greaterThan', formula=['0'], fill=green_fill))
ws_pf.conditional_formatting.add(f'K2:K{total_row-2}',
    CellIsRule(operator='lessThan', formula=['0'], fill=red_fill))
ws_pf.conditional_formatting.add(f'L2:L{total_row-2}',
    CellIsRule(operator='greaterThan', formula=['0'], fill=green_fill))
ws_pf.conditional_formatting.add(f'L2:L{total_row-2}',
    CellIsRule(operator='lessThan', formula=['0'], fill=red_fill))

# Cảnh báo khi giá đạt target (cột I vs M)
ws_pf.conditional_formatting.add(f'I2:I{total_row-2}',
    FormulaRule(formula=[f'I2>=M2'], fill=PatternFill('solid', start_color='FBBF24')))

ws_pf.freeze_panes = 'B2'

# =================== SHEET 3: TARGETS (Cảnh báo Telegram) ===================
ws_tg = wb.create_sheet('Alerts')

tg_headers = ['Token', 'Mạng', 'Loại Alert', 'Giá Trigger', 'Trạng Thái', 'Lần Cuối Check', 'Ghi Chú']
for col, header in enumerate(tg_headers, 1):
    cell = ws_tg.cell(row=1, column=col, value=header)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = THIN_BORDER

sample_alerts = [
    ['PEPE', 'ETH', 'TAKE_PROFIT', 0.00005, 'ACTIVE', '', 'Chốt lời x5'],
    ['PEPE', 'ETH', 'STOP_LOSS', 0.000005, 'ACTIVE', '', 'Cắt lỗ nếu xuống'],
    ['BRETT', 'BASE', 'TAKE_PROFIT', 0.5, 'ACTIVE', '', 'Target ATH'],
]
for row_idx, row_data in enumerate(sample_alerts, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_tg.cell(row=row_idx, column=col_idx, value=value)
        cell.font = INPUT_FONT
        cell.border = THIN_BORDER

widths_tg = [12, 8, 18, 18, 12, 18, 25]
for i, w in enumerate(widths_tg, 1):
    ws_tg.column_dimensions[get_column_letter(i)].width = w

ws_tg.freeze_panes = 'A2'

# =================== SHEET 4: DASHBOARD (Tổng quan) ===================
ws_db = wb.create_sheet('Dashboard', 0)  # Đặt làm sheet đầu tiên

ws_db['A1'] = '📊 DASHBOARD - QUẢN LÝ ĐẦU TƯ CRYPTO'
ws_db['A1'].font = Font(name='Arial', bold=True, size=16, color='1F2937')
ws_db.merge_cells('A1:F1')
ws_db['A1'].alignment = CENTER

# KPI cards
kpis = [
    ('💰 Tổng Vốn Đầu Tư', '=Portfolio!G5', '$#,##0.00'),
    ('📈 Giá Trị Hiện Tại', '=Portfolio!J5', '$#,##0.00'),
    ('💵 P&L Tổng', '=Portfolio!K5', '$#,##0.00;[Red]($#,##0.00)'),
    ('🎯 ROI %', '=Portfolio!L5', '0.00%;[Red](0.00%)'),
    ('🪙 Số Token Đang Giữ', '=COUNTA(Portfolio!A2:A50)-1', '0'),
    ('📝 Tổng Số Giao Dịch', '=COUNTA(Transactions!A2:A1000)', '0'),
]

for i, (label, formula, fmt) in enumerate(kpis):
    row = 3 + (i // 3) * 3
    col = (i % 3) * 2 + 1
    
    label_cell = ws_db.cell(row=row, column=col, value=label)
    label_cell.font = Font(name='Arial', bold=True, size=11, color='6B7280')
    label_cell.alignment = CENTER
    
    value_cell = ws_db.cell(row=row+1, column=col, value=formula)
    value_cell.font = Font(name='Arial', bold=True, size=18, color='1F2937')
    value_cell.alignment = CENTER
    value_cell.number_format = fmt
    
    # Merge 2 cột
    ws_db.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)
    ws_db.merge_cells(start_row=row+1, start_column=col, end_row=row+1, end_column=col+1)
    
    # Background
    for r in [row, row+1]:
        for c in [col, col+1]:
            try:
                ws_db.cell(row=r, column=c).fill = PatternFill('solid', start_color='F3F4F6')
            except:
                pass

# Khu vực phân bổ theo Network
ws_db['A11'] = '🌐 PHÂN BỔ THEO MẠNG'
ws_db['A11'].font = Font(name='Arial', bold=True, size=13)
ws_db.merge_cells('A11:F11')

ws_db['A12'] = 'Mạng'
ws_db['B12'] = 'Tổng Giá Trị USD'
ws_db['C12'] = '% Danh Mục'
for col_letter in ['A', 'B', 'C']:
    ws_db[f'{col_letter}12'].font = HEADER_FONT
    ws_db[f'{col_letter}12'].fill = HEADER_FILL
    ws_db[f'{col_letter}12'].alignment = CENTER

networks = ['ETH', 'BSC', 'BASE', 'SONEIUM', 'ARB', 'KATANA']
for i, net in enumerate(networks):
    row = 13 + i
    ws_db.cell(row=row, column=1, value=net).font = INPUT_FONT
    ws_db.cell(row=row, column=2, 
               value=f'=SUMIFS(Portfolio!J:J,Portfolio!B:B,A{row})').font = LINK_FONT
    ws_db.cell(row=row, column=2).number_format = '$#,##0.00'
    ws_db.cell(row=row, column=3, 
               value=f'=IFERROR(B{row}/SUM($B$13:$B$18),0)').font = FORMULA_FONT
    ws_db.cell(row=row, column=3).number_format = '0.00%'

# Column widths Dashboard
for col_letter in ['A', 'B', 'C', 'D', 'E', 'F']:
    ws_db.column_dimensions[col_letter].width = 22

ws_db.row_dimensions[1].height = 30
for r in [3, 6, 9]:
    ws_db.row_dimensions[r].height = 20
    ws_db.row_dimensions[r+1].height = 35

# =================== SHEET 5: HƯỚNG DẪN ===================
ws_help = wb.create_sheet('Hướng_Dẫn')

help_content = [
    ('📖 HƯỚNG DẪN SỬ DỤNG', True, 16),
    ('', False, 11),
    ('🎯 MỤC ĐÍCH', True, 13),
    ('App quản lý danh mục đầu tư crypto: theo dõi giá mua, tính P&L tự động, cảnh báo khi đạt target.', False, 11),
    ('', False, 11),
    ('📋 CÁC SHEET', True, 13),
    ('1. Dashboard: Xem tổng quan - tự động cập nhật từ các sheet khác', False, 11),
    ('2. Transactions: Nhập từng giao dịch mua/bán. Mỗi lần mua thêm = 1 dòng mới', False, 11),
    ('3. Portfolio: Tổng hợp tự động theo token - giá TB, P&L, % danh mục', False, 11),
    ('4. Alerts: Đặt cảnh báo Telegram khi giá chạm target', False, 11),
    ('', False, 11),
    ('✅ CÁCH NHẬP DỮ LIỆU', True, 13),
    ('• Ô MÀU XANH = bạn nhập tay', False, 11),
    ('• Ô MÀU ĐEN = công thức tự tính, KHÔNG sửa', False, 11),
    ('• Ô MÀU XANH LÁ = link sang sheet khác, KHÔNG sửa', False, 11),
    ('• Ô NỀN VÀNG = giá hiện tại do bot Python tự cập nhật', False, 11),
    ('', False, 11),
    ('💡 CÔNG THỨC GIÁ TRUNG BÌNH', True, 13),
    ('Công thức weighted average (trung bình có trọng số):', False, 11),
    ('Giá TB = TỔNG (Số lượng × Giá mua + Phí gas) ÷ TỔNG Số lượng', False, 11),
    ('Ví dụ: Mua lần 1: 1M PEPE @ $0.0000085 + phí $5.2 = $13.7', False, 11),
    ('       Mua lần 2: 500K PEPE @ $0.0000095 + phí $4.8 = $9.55', False, 11),
    ('       Giá TB = ($13.7 + $9.55) ÷ 1.5M = $0.0000155 / 1000 = $0.0000155', False, 11),
    ('', False, 11),
    ('🤖 BOT PYTHON', True, 13),
    ('1. Chạy: streamlit run app.py - mở UI để nhập giao dịch nhanh', False, 11),
    ('2. Chạy: python price_updater.py - update giá real-time mỗi 5 phút', False, 11),
    ('3. Chạy: python telegram_alert.py - bot cảnh báo khi giá đạt target', False, 11),
]

for row_idx, (text, bold, size) in enumerate(help_content, 1):
    cell = ws_help.cell(row=row_idx, column=1, value=text)
    cell.font = Font(name='Arial', bold=bold, size=size,
                     color='1F2937' if bold else '374151')
    cell.alignment = LEFT

ws_help.column_dimensions['A'].width = 100

# Save
wb.save('/home/claude/crypto_portfolio/Crypto_Portfolio_Template.xlsx')
print("✅ Template created!")
